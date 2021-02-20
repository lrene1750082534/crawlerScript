#!/usr/bin/python
# -*- coding: UTF-8 -*-
import re
import time
from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from difflib import SequenceMatcher


class Aiqiyi:
    def __init__(self):
        self.browser = webdriver.Chrome(r"D:\pro-code\local code\aiqiyi\config\chromedriver.exe")

    def close(self):
        '''
        关闭谷歌浏览器和驱动
        :return:
        '''
        self.browser.quit()
        return

    def verity(self, obj, name):
        '''
        校验爬取的名字和表格提供的是否一致
        :param obj:
        :param name:
        :return:
        '''
        if obj == name:
            return True
        else:
            return False

    def similarity(self, a, b):
        '''
        比较爬取的视频名和表格提供的名字相似度
        :param a:
        :param b:
        :return:
        '''
        return SequenceMatcher(None, a, b).ratio()

    def aliasnamecheck(self, temp, res):
        '''
        使用表格提供的又名爬取
        :param temp:
        :param res:
        :return:
        '''
        pattern = re.compile("又名")
        m = pattern.search(temp[0])
        if m:
            name = temp[0].split("：")
            if len(name) == 1:
                name = temp[0].split(":")
            print(name[1])
            item = re.findall(r'[^<>/h1第0-9页a-zA-Z- .]', name[1])
            # 正则去除^<>/h1第0-9页a-zA-Z . 这些符号
            item = ''.join(item)
            print("item:", item, "res:", res)
            el = self.verity(item, res)
            return el
        else:
            return False

    def isnameexist(self, chinesename):
        '''
        查看指定视频在平台是否存在
        :param chinesename:
        :return:
        '''
        try:
            self.browser.find_element_by_css_selector(
                "#__layout > div > div.qy20-page-wrap > div > div > div.qy-search-main > div.qy-search-result-con > div.layout-main > div:nth-child(2) > div:nth-child(2) > div.qy-search-result-item.vertical-pic > div.result-right > h3 > a")
            res = self.browser.find_element_by_xpath(
                '//*[@id="__layout"]/div/div[2]/div/div/div[2]/div[2]/div[1]/div[2]/div[2]/div[1]/div[2]/h3/a').get_attribute(
                "title")
            print("获取名字：", res)
            temp = re.findall(r"[（](.*?)[）]", chinesename)  # 括号内的内容，无则为空
            print(temp)
            if not temp:
                em = self.verity(res, chinesename)
                print("em:", em)
                return em
            else:
                rs = re.findall(r'([^（\）]+)(?:$|\（)', chinesename)  # 括号外的内容
                print("rs:", rs)
                rep = self.verity(res, rs[0])
                if not rep:
                    print("又名处理")
                    ret = self.aliasnamecheck(temp, res)
                    print("ret:", ret)
                    return ret
                else:
                    return rep
        except Exception as e:
            print("err:", e)
            return False

    def isenglishname(self, soup):
        '''
        获取指定视频是否提供英文名
        :param soup:
        :return:
        '''
        try:
            name = soup.select(".info-title-englishName span")
            print("name:", name)
            name = name[0].get("title")
            return name
        except Exception as e:
            print("error:", e)
            return None

    def get_name(self, name):
        '''
        执行任务
        :param name:
        :return:
        '''
        tag = False  # 标识视频是否能查到
        url = "https://www.iqiyi.com/"
        self.browser.get(url)
        self.browser.find_element_by_id("J-header-search-input").send_keys(name)
        self.browser.find_element_by_id("J-search-btn").click()
        n = self.browser.window_handles
        self.browser.switch_to.window(n[1])
        time.sleep(3)
        if self.isnameexist(name):
            try:
                tag = True
                print("123123123123123")
                self.browser.find_element_by_xpath(
                    '//*[@id="__layout"]/div/div[2]/div/div/div[2]/div[2]/div[1]/div[2]/div[2]/div[1]/div[2]/div[1]/div[5]/a').click()
                time.sleep(3)
                n = self.browser.window_handles
                self.browser.switch_to.window(n[2])
                currunt_url = self.browser.current_url
                print("current_url:", currunt_url)
                page_source = self.browser.page_source
                soup = BeautifulSoup(str(page_source), 'lxml')
                name = self.isenglishname(soup)
                print("othername:", name)
                return tag, name
            except Exception as e:
                print("getname err:", e)
                return tag, ""
            finally:
                self.browser.quit()
        else:
            self.close()
            return tag, ""


if __name__ == "__main__":
    wb = load_workbook(r"D:\pro-code\local code\aiqiyi\2021_02_09\香港亚视电视台-本地生产项目清单（2020-2015年内容）.xlsx")
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    ws['AN4'] = "爱奇艺"
    for i in range(5, ws.max_row + 1):
        y = 'AN' + str(i)
        name = ws[i][3].value
        eng = ws[i][2].value
        print('中文名-->{}'.format(name))
        print('英文名-->{}'.format(eng))
        e_name = Aiqiyi().get_name(name)
        print("最终结果：", e_name)
        try:
            if not e_name:
                e_name = None
            if e_name[0] and not eng or e_name[0] and ''.join(eng.split()).lower() == ''.join(
                    e_name[1].split()).lower():
                print("+++++++++++++")
                ws[y].value = "有"
            else:
                print("-----------")
                ws[y].value = "无"
        except Exception as e:
            pass
        time.sleep(3)
        print('============================================')
        wb.save(r"D:\pro-code\local code\aiqiyi\2021_02_09\香港亚视电视台-本地生产项目清单（2020-2015年内容）.xlsx")

# TODO:
#   1、爬取英文名和表格提供的比对
#   2、查看爱奇艺使用英文名查询中文名
#   3、整改结果显示形式
#   4、表格未提供英文名，爬取英文为空，记为有
#   5、表格提供英文名，爬取到英文名，两者比对
#   6、表格提供英文名，未爬取到英文名，记为无
