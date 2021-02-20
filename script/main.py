#!/usr/bin/python
# -*- coding: UTF-8 -*-
import re
import time
import os
import requests
import random
from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import load_workbook
from fake_useragent import UserAgent


class DouBan:
    def __init__(self):
        self.ips = self.test_proxy()
        self.driver = self.initdriver()

    def get_cookie(self, fileName=r'D:\pro-code\local code\aiqiyi\config\doucook.ini'):
        '''
        获取cookies值
        :param fileName:
        :return:
        '''
        ret = []
        if not os.path.exists(fileName):
            return ret
        with open(fileName, 'r', encoding='utf-8') as f:
            ret = f.read()
        cookie = {i.split('=')[0]: i.split('=')[1] for i in ret.strip().split(';')}
        print("cook:", cookie)
        return cookie

    def initdriver(self):
        '''
        创建谷歌驱动的实例
        :return:
        '''
        proxy = random.choice(self.ips)
        print("proxy=======:", proxy)

        option = webdriver.ChromeOptions()
        # webdriver.DesiredCapabilities.CHROME['proxy'] = {
        #     "httpProxy": proxy,
        #     "ftpProxy": proxy,
        #     "sslProxy": proxy,
        #     "proxyType": "MANUAL",
        #
        # }
        option.add_argument("--proxy-server=http://{}".format(proxy))
        webdriver.DesiredCapabilities.CHROME['acceptSslCerts'] = True
        # option.add_argument('--headless')
        option.add_argument('user-agent={}'.format(
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.104 Safari/537.36'))
        # option.add_argument('user-agent={}'.format(UserAgent().random))
        option.add_experimental_option('excludeSwitches', ['enable-automation'])
        driver = webdriver.Chrome(r"D:\pro-code\local code\aiqiyi\config\chromedriver.exe",
                                  options=option)
        return driver

    def test_proxy(self):
        '''
        获取IP列表
        :return:
        '''
        fp = open('../proxy_IP/host.txt', 'r')
        ips = fp.readlines()
        proxys = list()
        for p in ips:
            ip = p.strip('\n').split(':')
            proxy = ip[0] + ':' + ip[1]
            proxys.append(proxy)
        # proxys = ["182.96.51.4:4232","120.38.64.233:4245"]
        return proxys

    def aliasname(self, temp):
        '''
        获取表格提供的又名查询
        :param temp: 提供的视频又名
        :return:
        '''
        pattern = re.compile("又名")
        m = pattern.search(temp[0])
        if m:
            name = temp[0].split("：")
            print("name:", name)
            if len(name) == 1:
                name = temp[0].split(":")
            item = re.findall(r'[^<>/h1第0-9页a-zA-Z- .]', name[1])
            item = ''.join(item)
            print("item:", item)
            return item

    def dealname(self, name):
        '''
        处理表格提供的视频名
        :param name: 提供的视频名
        :return:
        '''
        temp = re.findall(r"[（](.*?)[）]", name)
        print(temp)
        if not temp:
            print("em:", name)
            return name, ""
        else:
            rs = re.findall(r'([^（\）]+)(?:$|\（)', name)
            print("rs:", rs)
            print("又名处理")
            ret = self.aliasname(temp)
            print("ret:", ret)
            return rs[0], ret

    def get_ename_2(self, href):
        '''
        获取平台提供的英文名
        :param href:
        :return:
        '''
        self.driver.get(href)
        time.sleep(2)
        text = str(self.driver.find_element_by_id('info').text)
        rows = text.split('\n')
        ename = ''
        for row in rows:
            if '又名' in row:
                enames = re.findall('[a-zA-Z]+', row)
                for item in enames:
                    ename = ename + ' ' + item
        if ename == '':
            print('没有找到英文名称')
        else:
            print('英文名称为：{}'.format(ename))
        return ename

    def get_ename(self, name, engname):
        '''
        执行任务
        :param name: 提供的视频中文名
        :param engname: 提供的视频英文名
        :return:
        '''
        ename = ''
        try:
            # 处理获取的名字
            name = self.dealname(name)
            url = 'https://search.douban.com/movie/subject_search?search_text={}&cat=1002'.format(name[0])
            self.driver.get(url)
            time.sleep(2)
            page_source = self.driver.page_source
            soup = BeautifulSoup(str(page_source), 'lxml')
            print("soup===============:", soup)
            root = soup.find('div', id='root')
            items = root.find_all('div', class_='item-root')

            if len(items) == 0:
                print('没有检索到该电影')
                if name[1]:
                    ename = self.get_ename(name[1], engname)
            elif len(items) == 1:
                item = items[0]
                soup = BeautifulSoup(str(item), 'lxml')
                title_text = str(soup.find('a', class_='title-text').text)[:-7]
                title_href = soup.find('a', class_='title-text')['href']
                print('查找到唯一结果：《{}》'.format(title_text))
                print('开始查找英文名称。。。')
                ename = self.get_ename_2(title_href)
                if not ename and name[1]:
                    print("======================?")
                    ename = self.get_ename(name[1], engname)
                    print("123:", ename)
            else:
                print('查找到多个相似结果，开始筛选。。。')
                for item in items:
                    soup = BeautifulSoup(str(item), 'lxml')
                    title_text = str(soup.find('a', class_='title-text').text[:-7])
                    title_href = soup.find('a', class_='title-text')['href']
                    if str(name[0]) in title_text:
                        print('查找到符合结果：《{}》'.format(title_text))
                        print('开始查找英文名称。。。')
                        ename = self.get_ename_2(title_href)
                        if not ename and name[1]:
                            ename = self.get_ename(name[1], engname)
                print('没有筛选到合适数据')
            print("endname:", engname)
            print("ename:", ename)
            if ename != engname:
                print("-------------------?")
                return False
            return ename
        except Exception as e:
            print("get_ename:", e)
            return ename

    def load_excel(self, path):
        '''
        加载表格，获取内容
        :param path: 表格地址
        :return:
        '''
        try:
            wb = load_workbook(path)
            sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            ws['AM4'] = "豆瓣"
            for i in range(5, ws.max_row + 1):
                y = 'AM' + str(i)
                name = ws[i][3].value  # 中文名
                engname = ws[i][2].value  # 英文名
                print('正在查找《{}》的英文名称。。。'.format(name))
                print('表格提供的英文名称《{}》。。。'.format(engname))
                e_name = self.get_ename(name, engname)
                print("最终结果：", e_name)
                if e_name:
                    ws[y].value = "有"
                else:
                    ws[y].value = "无"
                wb.save(path)
                time.sleep(2)
                print('============================================')
        except Exception as e:
            print("loadexcel error:", e)
        finally:
            self.driver.quit()


if __name__ == "__main__":
    path = r"D:\pro-code\local code\aiqiyi\2021_02_09\香港亚视电视台-本地生产项目清单（2020-2015年内容）.xlsx"
    e_name = DouBan().load_excel(path)
