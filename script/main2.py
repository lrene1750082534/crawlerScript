from selenium import webdriver
from openpyxl import load_workbook
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
import re
import time

option = webdriver.ChromeOptions()

option.add_argument('user-agent={}'.format(
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.104 Safari/537.36'))
driver = webdriver.Chrome(r"D:\pro-code\local code\aiqiyi\config\chromedriver.exe", options=option)


class DouBan:
    def __init__(self):
        pass

    def aliasname(self, temp):
        pattern = re.compile("又名")
        m = pattern.search(temp[0])
        if m:
            name = temp[0].split("：")
            # print("name:", name)
            if len(name) == 1:
                name = temp[0].split(":")
            item = re.findall(r'[^<>/h1第0-9页a-zA-Z- .]', name[1])
            item = ''.join(item)
            # print("item:", item)
            return item

    def dealname(self, name):
        temp = re.findall(r"[（](.*?)[）]", name)
        # print(temp)
        if not temp:
            # print("em:", name)
            return name, ""
        else:
            rs = re.findall(r'([^（\）]+)(?:$|\（)', name)
            # print("rs:", rs)
            print("又名处理")
            ret = self.aliasname(temp)
            # print("ret:", ret)
            return rs[0], ret

    def get_ename_2(self, href):
        driver.get(href)
        time.sleep(2)
        text = str(driver.find_element_by_id('info').text)
        rows = text.split('\n')
        ename = ''
        for row in rows:
            if '又名' in row:
                enames = re.findall('[a-zA-Z]+', row)
                for item in enames:
                    ename = ename + ' ' + item
        if ename == '':
            print('没有找到英文名称')
        else:
            print('英文名称为：{}'.format(ename))
        return ename

    def get_ename(self, name):
        ename = ''
        # 处理获取的名字
        name = self.dealname(name)
        url = 'https://search.douban.com/movie/subject_search?search_text={}&cat=1002'.format(name[0])
        driver.get(url)
        agent = driver.execute_script("return navigator.userAgent")
        print("selenium request header ===:", agent)
        time.sleep(2)
        page_source = driver.page_source
        soup = BeautifulSoup(str(page_source), 'lxml')
        root = soup.find('div', id='root')
        items = root.find_all('div', class_='item-root')
        if len(items) == 0:
            print('没有检索到该电影')
            if name[1]:
                self.get_ename(name[1])
            return ""
        elif len(items) == 1:
            item = items[0]
            soup = BeautifulSoup(str(item), 'lxml')
            title_text = str(soup.find('a', class_='title-text').text)[:-7]
            # print("title_text123:", str(soup.find('a', class_='title-text').text))
            # print("title_text:", title_text)
            title_href = soup.find('a', class_='title-text')['href']
            print('查找到唯一结果：《{}》'.format(title_text))
            print('开始查找英文名称。。。')
            ename = self.get_ename_2(title_href)
            if not ename and name[1]:
                self.get_ename(name[1])
            return ename
        else:
            print('查找到多个相似结果，开始筛选。。。')
            for item in items:
                soup = BeautifulSoup(str(item), 'lxml')
                title_text = str(soup.find('a', class_='title-text').text[:-7])
                title_href = soup.find('a', class_='title-text')['href']
                if str(name[0]) in title_text:
                    print('查找到符合结果：《{}》'.format(title_text))
                    print('开始查找英文名称。。。')
                    ename = self.get_ename_2(title_href)
                    if not ename and name[1]:
                        self.get_ename(name[1])
                    return ename
            print('没有筛选到合适数据')
        return ename

    def load_excel(self, path):
        wb = load_workbook(path)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        for i in range(5, ws.max_row + 1):
            # name = ws[i][3].value
            name = "吉祥任务（又名: 异度任务）"
            print('正在查找《{}》的英文名称。。。'.format(name))
            e_name = self.get_ename(name)
            try:
                ws[i][2].value = e_name
            except Exception as e:
                pass
            time.sleep(2)
            print('============================================')
        wb.save(path)


if __name__ == "__main__":
    path = "./香港亚视电视台-本地生产项目清单（2020-2015年内容）_41.xlsx"
    e_name = DouBan().load_excel(path)
