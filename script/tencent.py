#!/usr/bin/python
# -*- coding: UTF-8 -*-
import re
import time
from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from difflib import SequenceMatcher


class Tencent:
    def __init__(self):
        option = webdriver.ChromeOptions()
        option.add_argument('--headless')
        self.browser = webdriver.Chrome(r"D:\pro-code\local code\aiqiyi\config\chromedriver.exe", options=option)

    def close(self):
        '''
        关闭浏览器
        :return:
        '''
        self.browser.quit()
        return

    def similarity(self, a, b):
        '''
        比对搜索结果视频名与指定视频名相似度
        相似度60%及以上返回True,否则返回False
        :param a: 搜索结果名
        :param b: 指定视频名
        :return:
        '''
        if SequenceMatcher(None, a, b).ratio() >= 0.6:
            print("相似度：", SequenceMatcher(None, a, b).ratio())
            return True
        else:
            return False

    def pretreatment(self, name):
        '''
        获取提供视频的原名 & 又名
        :param name: 提供的视频名
        :return: 原名，又名 or 原名，""
        '''
        # 取括号以外的内容
        mainname = re.findall(r'([^（\）]+)(?:$|\（)', name)[0]
        temp = re.findall(r"[（](.*?)[）]", name)
        if temp:
            pattern = re.compile("又名")
            m = pattern.search(temp[0])
            if m:
                name = temp[0].split("：")
                print(name[1])
                item = re.findall(r'[^<>/h1第0-9页a-zA-Z- .]', name[1])
                # 正则去除^<>/h1第0-9页a-zA-Z . 这些符号
                item = ''.join(item)
                return mainname, item
        return mainname, ""

    def deepsearch(self, items):
        '''
        页首匹配失败，向下匹配
        :param items:
        :return:
        '''
        for item in items:
            soup2 = BeautifulSoup(str(item), 'lxml')
            print("item", "=" * 20, soup2)
            first_title = str(soup2.find('h2', class_='result_title'))
            soup3 = BeautifulSoup(str(first_title), 'lxml')
            title = soup3.em.text
            print("title:", title)
            # 比对标题和搜索的是否相似
            if self.similarity(first_title, name):
                print('查找到符合结果：《{}》'.format(first_title))
                return True, soup2
            else:
                # 比对展示的别名和搜索的是否相似
                othername = str(soup2.find('span', class_='label').get_text).strip()
                if self.similarity(othername, name):
                    return True, soup2
        print('没有筛选到合适数据')
        return False

    def hasresult(self, name):
        '''
        判断是否有搜索结果
        :param name: 指定视频名
        :return: 有结果返True，
                 无结果返False
        '''
        try:
            page_source = self.browser.page_source
            soup = BeautifulSoup(str(page_source), 'lxml')
            root = soup.find('div', class_='wrapper_main')
            items = root.find_all('div', class_='result_item result_item_v')
            print("items:", items)
            if len(items) == 0:
                print("no result")
                return False
            else:
                return self.deepsearch(items)
        except Exception as e:
            print("hasresult,e:", e)
            return False

    def hasename(self, soup):
        '''
        判断该视频是否提供英文名
        :return: 提供返回英文名，
                 不提供返回False
        '''
        try:
            items = soup.find('div', class_='info_item info_item_odd')
            if len(items) == 0:
                print("没数据")
            else:
                for item in items:
                    soup2 = BeautifulSoup(str(item), 'lxml')
                    onename = str(soup2.find('span', class_='label').text).strip()
                    if onename == "英文名：":
                        ename = str(soup2.find('span', class_='content').text).strip()
                        print("ename:", ename)
                        return ename
            return ""
        except Exception as e:
            print("getname e:", e)
            return ""

    def work(self, name):
        '''
        爬取内容
        :param name: 搜索视频名
        :return: 搜索视频对应的英文名，没有则返空字符串
        '''
        try:
            url = "https://v.qq.com/x/search/?q={}&stag=0&smartbox_ab=".format(name)
            self.browser.get(url)
            # 判断是否有搜索的视频
            res = self.hasresult(name)
            if res[0]:
                # 判断该视频是否提供英文名
                engname = self.hasename(res[1])
                if engname:
                    return engname
            return ""
        except Exception as e:
            print("err:", e)
            return ""

    def main(self, name):
        '''
        主函数
        :param name:
        :return:
        '''
        tag = True
        try:
            pre = self.pretreatment(name)
            res = self.work(pre[0])
            # 原名没查到，用又名
            print("pre:", pre)
            print("res:", res)
            if not res and pre[1]:
                res = self.work(pre[1])
                if not res:
                    tag = False
            else:
                tag = False

            return tag, res
        except Exception as e:
            print("main error:", e)
            return False, ""
        finally:
            self.close()


if __name__ == "__main__":
    wb = load_workbook(r"D:\pro-code\local code\aiqiyi\2021_02_09\香港亚视电视台-本地生产项目清单（2020-2015年内容）.xlsx")
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    ws["AO4"] = "腾讯"
    for i in range(5, ws.max_row + 1):
        y = 'AO' + str(i)
        name = ws[i][3].value
        eng = ws[i][2].value
        print('中文名-->{}'.format(name))
        print('英文名-->{}'.format(eng))
        e_name = Tencent().main(name)
        print("最终结果：", e_name)
        try:
            if e_name[0] and not eng or e_name[0] and ''.join(eng.split()).lower() == ''.join(
                    e_name[1].split()).lower():
                print("+++++++++++++")
                ws[y].value = "有"
            else:
                print("-----------")
                ws[y].value = "无"

        except Exception as e:
            pass
        time.sleep(3)
        print('============================================')
        wb.save(r"D:\pro-code\local code\aiqiyi\2021_02_09\香港亚视电视台-本地生产项目清单（2020-2015年内容）.xlsx")

# TODO :
#   1、视频存在，表格提供英文名：
#       a.爬取到英文名-->比对，一致，有；否则无；
#       b.未爬取到英文名-->无
#   2、整改结果表现形式
#   3、中文名查不到用提供的英文名查
