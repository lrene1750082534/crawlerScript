#!/usr/bin/python
# -*- coding: UTF-8 -*-
import re
import time
import requests
from urllib import parse
from bs4 import BeautifulSoup
from openpyxl import load_workbook


class Youku:
    def __init__(self):
        pass

    def aliasnamecheck(self, temp, title):
        '''
        使用表格提供的又名爬取
        :param temp:
        :param res:
        :return:
        '''
        pattern = re.compile("又名")
        m = pattern.search(temp)
        if m:
            name = temp.split("：")
            if len(name) == 1:
                name = temp.split(":")
            print(name[1])
            if name[1] == title:
                return True
        return False

    def isnameexist(self, chinesename, title):
        '''
        查看指定视频在平台是否存在
        :param chinesename:
        :return:
        '''
        try:
            temp = re.findall(r"[（](.*?)[）]", chinesename)  # 括号内的内容
            if not temp and chinesename == title:
                return True
            else:
                rs = re.findall(r'([^（\）]+)(?:$|\（)', chinesename)  # 括号外的内容
                if rs[0] == title:
                    return True
                else:
                    result = self.aliasnamecheck(temp[0], title)
                    return result
        except Exception as e:
            print("err:", e)
            return False

    def send_request(self, name):
        '''
        发送请求，获取渲染的页面数据
        :param name: 表格提供的视频中文名
        :return:
        '''
        vedioname = parse.quote(name)
        url = 'https://so.youku.com/search_video/q_{}?searchfrom=1'.format(vedioname)
        headers = {
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
            "accept-encoding": "gzip, deflate, br",
            "accept-language": "zh-CN,zh;q=0.9,en;q=0.8",
            "cache-control": "no-cache",
            "cookie": "_uab_collina=161361364927426107739567; cna=VYYWGOpN5gkCAXL0JIK3YDxv; __ysuid=16136136493369K9; __ayft=1613613649343; __aysid=1613613649344kdU; __ayscnt=1; xlly_s=1; ctoken=hQRokx_V0xfOb57BjnG2PAbG; _m_h5_tk=91b084f01a2d48598a29403d0cc23609_1613618167025; _m_h5_tk_enc=1c08c66ff9f68499bbb5aaf7b51278d9; UM_distinctid=177b2ea6f0a18-0a0c404a0b4452-53e3566-144000-177b2ea6f0b6a8; P_sck=V7wU57eViOUGtaWmBkAvurEtyjsV%2Fgm8QPzr3yUw%2FXpq6RmwScwxLb6Sh5CKF2spCZxzt7LBFVyEhSUbkE1RUSz9k7odBvhxMdKHJ8Wz6NWOXIpCSfeFfB91W5pagsq79okpFvPVvpYGb0Vt%2FXL5OQ%3D%3D; P_gck=NA%7CeF5tR7nBrs3VLiRmZZVmZQ%3D%3D%7CNA%7C1613614955981; P_pck_rm=02aVb%2FmP36aff7942a70b8ZBickxjod81tOXwasSx25%2BmYvJiuvns5P2cMEbmVrTBsdHHSjDfrN1Q%2B3xWmo%2FG6npK3bVqnt%2BIrqoXxkcEpWZuJKM0uGK%2BD%2BKDBQieY0ZuEAVNM9GDnK4FQ0IjSnpsgKWbUR7GObr8mHoRQ%3D%3D%5FV2; youku_history_word=%5B%22%25E7%2594%25B5%25E8%25A7%2586%25E9%25A3%258E%25E4%25BA%2591%22%5D; P_ck_ctl=16F522CEFDDAA0A6BE2FEF392CE94B4B; __arpvid=1613617811232tp3ljJ-1613617811259; __aypstp=18; __ayspstp=18; isg=BAkJZq9qjZUNYk675KNcrSLQGDVjVv2I5Qe3xqt-AvAv8ikE86dqWPVqMFbErZXA; tfstk=cctlBANwmU7Sc3bD50sWQvZn0I3Oao0PNH-20nhRLkZ9A2RlUsDbuPWtgxfJZHUC.; l=eBSA5GF7Ov9olzPoBO5aFurza77tmIRb41PzaNbMiInca6sNtFMCBNCIAUDWSdtjgt1hpety-Z76RdLHR3fCOxDDBvRlN5Ftnxf..",
            "pragma": "no-cache",
            "sec-ch-ua": '"Chromium";v="88", "Google Chrome";v="88", ";Not A Brand";v="99"',
            "sec-ch-ua-mobile": "?0",
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "none",
            "sec-fetch-user": "?1",
            "upgrade-insecure-requests": "1",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36"
        }
        response = requests.get(url=url, headers=headers)
        page_source = response.text
        con = response.content
        return page_source, con

    def main(self, name, year):
        '''
        主函数
        :param name: 表格提供的视频中文名
        :param year: 表格提供的视频年份
        :return:
        '''
        page_source = self.send_request(name)
        soup = BeautifulSoup(page_source[0], features="lxml")
        items = soup.find_all(attrs={"type": "1027"})
        if not items:
            print("无匹配结果1")
            with open(file="{}.html".format(name), mode="wb") as f:
                f.write(page_source[1])
            return False
        for item in items:
            titleobj = item.find('mark', class_='temp-hight-title_1ISDT')
            if titleobj:
                title = titleobj.text
                verity_title = self.isnameexist(name, title)
                if verity_title:
                    astr = item.find('div', class_='poster-desc_uTT1A').text
                    verify_year = astr.split(r'·')[1].strip()
                    if int(verify_year) == year:
                        print("success")
                        return True
        print("无匹配结果2")
        with open(file="{}.html".format(name), mode="wb") as f:
            f.write(page_source[1])
        print("------------------------")
        return False

    def load_excel(self, path):
        '''
        加载表格，获取信息
        :param path: 表格路径
        :return:
        '''
        wb = load_workbook(path)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        ws['AP4'] = "优酷"
        count = 1
        for i in range(5, ws.max_row + 1):
            print('>' * 20, count, '<' * 20)
            y = 'AP' + str(i)
            name = ws[i][3].value  # 中文名
            year = ws[i][0].value  # 年份
            print('正在查找《{}》的英文名称。。。'.format(name))
            print('表格提供的年份《{}》。。。'.format(year))
            res = self.main(name, year)
            print("最终结果：", res)
            if res:
                ws[y].value = "有"
            else:
                ws[y].value = "无"
            wb.save(path)
            count += 1
            time.sleep(2)
            print("===============================================================")


if __name__ == '__main__':
    # path = r"D:\pro-code\local code\aiqiyi\2021_02_09\香港亚视电视台-本地生产项目清单（2020-2015年内容）.xlsx"
    # e_name = Youku().load_excel(path)
    e_name = Youku().main("亲密损友（外判剧集）", 2012)
    print(">>>>>>>>>>>>:", e_name)
