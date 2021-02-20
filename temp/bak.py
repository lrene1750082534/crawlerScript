from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time
import re


def verity(obj, name):
    if obj == name:
        return True
    else:
        return False


def aliasnamecheck(temp, res):
    pattern = re.compile("又名")
    m = pattern.search(temp[0])
    if m:
        name = temp[0].split("：")
        print(name[1])
        item = re.findall(r'[^<>/h1第0-9页a-zA-Z- .]', name[1])
        # 正则去除^<>/h1第0-9页a-zA-Z . 这些符号
        item = ''.join(item)
        print("item:", item, "res:", res)
        el = verity(item, res)
        return el


def isnameexist(browser, chinesename):
    try:
        browser.find_element_by_css_selector(
            "#__layout > div > div.qy20-page-wrap > div > div > div.qy-search-main > div.qy-search-result-con > div.layout-main > div:nth-child(2) > div:nth-child(2) > div.qy-search-result-item.vertical-pic > div.result-right > h3 > a")
        res = browser.find_element_by_xpath(
            '//*[@id="__layout"]/div/div[2]/div/div/div[2]/div[2]/div[1]/div[2]/div[2]/div[1]/div[2]/h3/a').get_attribute(
            "title")
        print("获取名字：", res)
        temp = re.findall(r"[（](.*?)[）]", chinesename)
        print(temp)
        if not temp:
            em = verity(res, chinesename)
            print("em:", em)
            return em

        else:
            rs = re.findall(r'([^（\）]+)(?:$|\（)', chinesename)
            print("rs:", rs)
            if not verity(res, rs[0]):
                print("又名处理")
                ret = aliasnamecheck(temp, res)
                print("ret:", ret)
                return ret
    except Exception as e:
        print("err:", e)
        return False


def isenglishname(soup):
    try:
        name = soup.select(".info-title-englishName span")
        print("name:", name)
        name = name[0].get("title")
        return name
    except Exception as e:
        print("error:", e)
        return ""


def get_name(name):
    browser = webdriver.Chrome(r"D:\pro-code\local code\aiqiyi\config\chromedriver.exe")
    url = "https://www.iqiyi.com/"
    browser.get(url)
    browser.find_element_by_id("J-header-search-input").send_keys(name)
    browser.find_element_by_id("J-search-btn").click()
    n = browser.window_handles  # 获取当前页句柄
    browser.switch_to.window(n[1])  # 切换到新的网页窗口
    time.sleep(3)
    if isnameexist(browser, name):
        print("123123123123123")
        browser.find_element_by_xpath(
            '//*[@id="__layout"]/div/div[2]/div/div/div[2]/div[2]/div[1]/div[2]/div[2]/div[1]/div[2]/div[1]/div[5]/a').click()
        time.sleep(3)
        n = browser.window_handles
        browser.switch_to.window(n[2])
        currunt_url = browser.current_url
        print("current_url:", currunt_url)
        page_source = browser.page_source
        soup = BeautifulSoup(str(page_source), 'lxml')
        name = isenglishname(soup)
        print("othername:", name)
        browser.quit()
        return name
    else:
        browser.quit()
        return ""


wb = load_workbook("./香港亚视电视台-本地生产项目清单（2020-2015年内容）_41.xlsx")
sheet_name = wb.sheetnames[0]
ws = wb[sheet_name]
for i in range(5, ws.max_row + 1):
    name = ws[i][3].value
    print('正在查找《{}》的英文名称。。。'.format(name))
    # e_name = get_name("祖先开眼（合拍剧集）")
    e_name = get_name(name)
    try:
        ws[i][2].value = e_name
    except Exception as e:
        pass
    time.sleep(3)
    print('============================================')
wb.save("./香港亚视电视台-本地生产项目清单（2020-2015年内容）_41.xlsx")



